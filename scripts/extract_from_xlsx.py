"""
extract_from_xlsx.py
--------------------
Extract swim-club data from the source Excel workbook and write clean CSVs.

The source workbook contains helper columns, generated INSERT statements,
and formula residue alongside the actual data. This script keeps only the
actual data columns, applies silent fixes (ID normalisation, type coercion,
date formatting), and emits one CSV per entity ready for SQLite COPY.

Silent fixes applied during extraction:
  1. Nomination swimmerIDs normalised from 'swmr1' -> 'swmr01'
  2. Swimmer.age coerced from string '13' to integer 13
  3. Dates formatted as ISO 8601 (YYYY-MM-DD) for SQLite compatibility
  4. Time fields formatted as HH:MM:SS strings
  5. Empty lanes (null swimmer/result) preserved as NULL, not dropped

Usage:
    python extract_from_xlsx.py <source.xlsx> <output_dir>
"""

import csv
import sys
from datetime import date, datetime, time
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalise_swmr_id(value):
    """Normalise 'swmr1' -> 'swmr01'. Leaves already-padded IDs unchanged."""
    if value is None:
        return None
    s = str(value).strip()
    if s.startswith("swmr"):
        num = s.replace("swmr", "")
        if num.isdigit():
            return f"swmr{int(num):02d}"
    return s


def fmt_date(v):
    """Format a datetime/date as ISO 8601 date string, or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def fmt_time(v):
    """Format a time/timedelta as HH:MM:SS string, or None."""
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime):
        return v.strftime("%H:%M:%S")
    # Fallback: already a string
    return str(v)


def to_int(v):
    if v is None or v == "":
        return None
    return int(float(v))  # handles '13' and 13.0


def to_float(v):
    if v is None or v == "":
        return None
    return float(v)


def write_csv(path, header, rows):
    """Write rows to a CSV, using empty string for None."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            w.writerow(["" if v is None else v for v in r])
    print(f"  wrote {path.name:40s} {len(rows)} rows")


# ---------------------------------------------------------------------------
# Per-entity extractors
# ---------------------------------------------------------------------------

def extract_officials(wb, out):
    ws = wb["Official_done"]
    rows = []
    # header at row 2, data from row 3
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            rows.append((row[0], row[1], row[2], fmt_date(row[3]), row[4]))
    write_csv(out / "officials.csv",
              ["officialID", "firstName", "lastName", "employeeStartDate", "activeStatus"],
              rows)


def extract_coaches(wb, out):
    ws = wb["Coach_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], row[1], row[2], fmt_date(row[3])))
    write_csv(out / "coaches.csv",
              ["coachID", "firstName", "lastName", "employeeStartDate"],
              rows)


def extract_role_types(wb, out):
    ws = wb["RoleType_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], row[1], row[2], to_int(row[3])))
    write_csv(out / "role_types.csv",
              ["roleTypeID", "roleTitle", "roleDescription", "numberRequired"],
              rows)


def extract_racemeets(wb, out):
    ws = wb["RacemeetInstance_Done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], to_int(row[1]), fmt_date(row[2]), fmt_time(row[3])))
    write_csv(out / "racemeet_instances.csv",
              ["racemeetInstanceID", "weekOfCompetition", "racemeetDate", "startTime"],
              rows)


def extract_event_types(wb, out):
    ws = wb["EventType_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], row[1], row[2], to_int(row[3]), to_int(row[4]), row[5]))
    write_csv(out / "event_types.csv",
              ["eventTypeID", "eventDescription", "strokeType",
               "distancePerLeg", "numberOfLegs", "individualOrMedley"],
              rows)


def extract_lane_types(wb, out):
    ws = wb["LaneType_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], to_int(row[1])))
    write_csv(out / "lane_types.csv", ["laneTypeID", "laneNumber"], rows)


def extract_result_types(wb, out):
    ws = wb["ResultType_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], row[1], to_int(row[2])))
    write_csv(out / "result_types.csv",
              ["resultTypeID", "positionResult", "points"], rows)


def extract_divisions(wb, out):
    ws = wb["Division_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], to_int(row[1]), row[2], row[3]))
    write_csv(out / "divisions.csv",
              ["divisionID", "age", "gender", "coachID"], rows)


def extract_swimmers(wb, out):
    ws = wb["Swimmer_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            # age stored as string in source -> cast to int
            rows.append((row[0], row[1], row[2], fmt_date(row[3]),
                         to_int(row[4]), row[5], row[6]))
    write_csv(out / "swimmers.csv",
              ["swimmerID", "firstName", "lastName", "dob",
               "age", "gender", "divisionID"], rows)


def extract_racemeet_events(wb, out):
    ws = wb["RacemeetEventInstance_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], fmt_time(row[1]), row[2], row[3]))
    write_csv(out / "racemeet_event_instances.csv",
              ["racemeetEventInstanceID", "startTime",
               "racemeetInstanceID", "eventTypeID"], rows)


def extract_race_instances(wb, out):
    ws = wb["RaceInstance_done"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], fmt_time(row[1]), row[2], row[3], row[4]))
    write_csv(out / "race_instances.csv",
              ["raceInstanceID", "startTime", "poolLocation",
               "divisionID", "racemeetEventInstanceID"], rows)


def extract_race_lane_instances(wb, out):
    """RaceLaneInstance has headers at row 3, data from row 4.
    Columns 0..5 = raceLaneInstanceID, timeTakenToCompleteRace,
    raceInstanceID, laneTypeID, swimmerID, resultTypeID.

    Silent fix: the source spreadsheet contains a data-entry error where
    swimmer 'swmr85' is entered in BOTH lane01 and lane06 of every backstroke
    race for the 13M division (4 occurrences total). Cross-checking the
    nomination table confirms swmr85 never nominated for backstroke. We keep
    the first occurrence and null out the swimmerID/resultTypeID on the
    duplicate, preserving the lane slot rather than dropping the row.
    Documented in docs/design-notes.md.
    """
    ws = wb["RaceLaneInstance"]
    raw = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0]:
            raw.append([row[0], to_float(row[1]), row[2], row[3], row[4], row[5]])

    # Deduplicate (raceInstanceID, swimmerID) pairs where swimmerID is set.
    seen = set()
    fixed = 0
    for r in raw:
        race_id, swimmer_id = r[2], r[4]
        if swimmer_id:
            key = (race_id, swimmer_id)
            if key in seen:
                # null out the duplicate swimmer + result, keep time + lane
                r[4] = None
                r[5] = None
                fixed += 1
            else:
                seen.add(key)
    if fixed:
        print(f"  (applied silent fix: {fixed} duplicate swimmer-in-race entries nulled)")

    write_csv(out / "race_lane_instances.csv",
              ["raceLaneInstanceID", "timeTakenToCompleteRace",
               "raceInstanceID", "laneTypeID", "swimmerID", "resultTypeID"], raw)


def extract_official_roles(wb, out):
    ws = wb["OfficialRacemeetRoleInstance_do"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append((row[0], row[1], row[2]))
    write_csv(out / "official_racemeet_roles.csv",
              ["officialID", "racemeetInstanceID", "roleTypeID"], rows)


def extract_nominations(wb, out):
    """Apply silent fixes:
      1. Normalise swmr1 -> swmr01
      2. Dedupe on (swimmerID, racemeetEventInstanceID) - source data contains
         one such duplicate (nom46 == nom23 for swmr21 in rmev001).
    """
    ws = wb["NominationInstance"]
    raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            raw.append((row[0], fmt_date(row[1]), row[2], normalise_swmr_id(row[3])))

    seen = set()
    deduped = []
    for r in raw:
        key = (r[3], r[2])  # (swimmerID, racemeetEventInstanceID)
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    dropped = len(raw) - len(deduped)
    if dropped:
        print(f"  (applied silent fix: {dropped} duplicate nominations dropped)")

    write_csv(out / "nomination_instances.csv",
              ["nominationInstanceID", "dateNominated",
               "racemeetEventInstanceID", "swimmerID"], deduped)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 3:
        print("Usage: python extract_from_xlsx.py <source.xlsx> <output_dir>")
        sys.exit(1)

    src = Path(sys.argv[1])
    out = Path(sys.argv[2])
    out.mkdir(parents=True, exist_ok=True)

    print(f"Loading {src}...")
    wb = openpyxl.load_workbook(src, data_only=True)

    print("Extracting CSVs:")
    extract_officials(wb, out)
    extract_coaches(wb, out)
    extract_role_types(wb, out)
    extract_racemeets(wb, out)
    extract_event_types(wb, out)
    extract_lane_types(wb, out)
    extract_result_types(wb, out)
    extract_divisions(wb, out)
    extract_swimmers(wb, out)
    extract_racemeet_events(wb, out)
    extract_race_instances(wb, out)
    extract_race_lane_instances(wb, out)
    extract_official_roles(wb, out)
    extract_nominations(wb, out)
    print("Done.")


if __name__ == "__main__":
    main()
